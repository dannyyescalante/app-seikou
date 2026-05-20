#!/usr/bin/env python3
"""
Conciliación Bancaria – Seikou S.A.
Versión 2.0
Bancos: Bancolombia · BBVA · Colpatria · Davivienda · Occidente · IRIS · Finandina
"""

import streamlit as st
import pandas as pd
import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io, re
from datetime import datetime, date
import warnings
warnings.filterwarnings("ignore")

# ==============================================================
# CONFIGURACIÓN UI
# ==============================================================
st.set_page_config(
    page_title="Conciliación Bancaria – Seikou SA",
    page_icon="",
    layout="wide",
    initial_sidebar_state="expanded",
)

BANCOS = ["Bancolombia", "BBVA", "Colpatria", "Davivienda", "Occidente", "IRIS", "Finandina"]

# Palabras clave para identificar GMF / comisiones bancarias
GMF_KEYWORDS = [
    "4X1000", "4X1.000", "IMPTO GOBIERNO", "GMF", "GRAVAMEN",
    "COBRO COMISION QR", "COBRO IVA COMISION", "IVA CUOTA MANEJO",
    "IVA COBRADO", "COMISION PG TER", "COMISION ADMON",
    "CUOTA MANEJO", "IVA POR COMISION", "COBRO ADEUDO COM",
    "COBRO ADEUDO IVA", "CARGO POR IMPUESTO", "IVA POR SERVICIOS",
    "COBRO IVA PAGOS", "SERVICIO PAGO A OTROS BANCOS",
    "COMISION RECAUDO", "IVA COMISION RECAUDO",
]

# Cuentas bancarias por banco (para la carátula)
INFO_BANCO = {
    "Bancolombia": {"cuenta": "11100555", "num_cta": "No. 638-401981-73"},
    "BBVA":        {"cuenta": "11200520", "num_cta": "No. 001308330200018146"},
    "Colpatria":   {"cuenta": "11100525", "num_cta": "No. 012108616-6"},
    "Davivienda":  {"cuenta": "11100560", "num_cta": "No. 4776-6999-5870"},
    "Occidente":   {"cuenta": "11100571", "num_cta": "No. 261-87472-1"},
    "IRIS":        {"cuenta": "11200516", "num_cta": "No. 100739047057"},
    "Finandina":   {"cuenta": "",          "num_cta": ""},
}

# ==============================================================
# UTILIDADES NUMÉRICAS
# ==============================================================

def _clean_num(s):
    if s is None: return ""
    return str(s).strip().replace("$", "").replace("\xa0", "").replace(" ", "")

def parse_anglosajón(s):
    """coma=miles, punto=decimal  (Bancolombia, BBVA, Davivienda, Occidente, IRIS)"""
    v = _clean_num(s)
    if v in ("", "-", "-", "nan", "None", "0.0"): return 0.0
    neg = v.startswith("-"); v = v.lstrip("+-")
    try: return -float(v.replace(",", "")) if neg else float(v.replace(",", ""))
    except: return 0.0

def parse_europeo(s):
    """punto=miles, coma=decimal  (Colpatria)"""
    v = _clean_num(s)
    if v in ("", "-", "-", "nan", "None"): return 0.0
    neg = v.startswith("-"); v = v.lstrip("+-")
    try: return -float(v.replace(".", "").replace(",", ".")) if neg else float(v.replace(".", "").replace(",", "."))
    except: return 0.0

def is_gmf(desc: str) -> bool:
    d = str(desc).upper()
    return any(k in d for k in GMF_KEYWORDS)

def _empty_df():
    return pd.DataFrame(columns=["FECHA", "DESCRIPCION", "VALOR", "TIPO", "GMF"])

# ==============================================================
# HELPERS PDF
# ==============================================================

def _lines(page, y_tol=4):
    """Agrupa palabras por línea."""
    words = page.extract_words(x_tolerance=3, y_tolerance=y_tol)
    groups = {}
    for w in words:
        y = round(float(w["top"]) / y_tol) * y_tol
        groups.setdefault(y, []).append(w)
    return {y: sorted(v, key=lambda w: w["x0"]) for y, v in sorted(groups.items())}

def _txt(wlist): return " ".join(w["text"] for w in wlist)
def _at(wlist, x0, x1): return "".join(w["text"] for w in wlist if x0 <= w["x0"] < x1)
def _match(text, pat): return bool(re.match(pat, text.strip()))

def _find_header_xs(lines_dict, *keywords):
    """Busca posición x de columnas por su texto de encabezado."""
    result = {}
    for y, wlist in lines_dict.items():
        txt = _txt(wlist).upper()
        if all(k.upper() in txt for k in keywords):
            for w in wlist:
                for k in keywords:
                    if k.upper() in w["text"].upper():
                        result[k] = w["x0"]
            break
    return result

# ==============================================================
# EXTRACTORES PDF  (uno por banco)
# ==============================================================

# -- BANCOLOMBIA ---------------------------------------------
# Estructura: FECHA | DESCRIPCIÓN | SUCURSAL | DCTO. | VALOR | SALDO
# Fechas: d/mm  (ej. 1/03, 31/03)
# Números: anglosajón  – el VALOR ya trae signo (neg=débito, pos=crédito)

def extract_bancolombia(pdf_file, pdf_password=""):
    rows = []
    with pdfplumber.open(pdf_file, password=pdf_password or None) as pdf:
        for page in pdf.pages:
            lns = _lines(page)
            valor_x = saldo_x = None
            for y, wl in lns.items():
                t = _txt(wl).upper()
                if "VALOR" in t and "SALDO" in t:
                    for w in wl:
                        if "VALOR" in w["text"].upper(): valor_x = w["x0"]
                        if "SALDO" in w["text"].upper(): saldo_x = w["x0"]
                    break
            if valor_x is None: valor_x, saldo_x = 400, 470

            for y, wl in lns.items():
                if not wl: continue
                first = wl[0]["text"]
                if not _match(first, r"^\d{1,2}/\d{2}$"): continue

                desc_w = [w for w in wl if 40 <= w["x0"] < valor_x - 4]
                val_w  = [w for w in wl if valor_x - 4 <= w["x0"] < saldo_x - 4]

                fecha = first
                desc  = _txt(desc_w).strip()
                vstr  = "".join(w["text"] for w in val_w).strip()
                if not vstr: continue

                valor = parse_anglosajón(vstr)
                if valor == 0: continue
                rows.append({"FECHA": fecha, "DESCRIPCION": desc,
                              "VALOR": valor, "TIPO": "CREDITO" if valor > 0 else "DEBITO",
                              "GMF": is_gmf(desc)})
    return pd.DataFrame(rows) if rows else _empty_df()

# -- BBVA -----------------------------------------------------
# Estructura: Movi | FechaOp | FechaVal | Concepto | Cargos | Abonos | Saldo
# Línea TX: empieza con número de movimiento de 4 dígitos

def extract_bbva(pdf_file, pdf_password=""):
    rows = []
    with pdfplumber.open(pdf_file, password=pdf_password or None) as pdf:
        for page in pdf.pages:
            lns = _lines(page)
            cargo_x = abono_x = None
            for y, wl in lns.items():
                t = _txt(wl).upper()
                if "CARGOS" in t and "ABONOS" in t:
                    for w in wl:
                        if "CARGO" in w["text"].upper(): cargo_x = w["x0"]
                        if "ABONO" in w["text"].upper(): abono_x = w["x0"]
                    break
            if cargo_x is None: cargo_x, abono_x = 420, 500

            for y, wl in lns.items():
                if not wl: continue
                if not _match(wl[0]["text"], r"^\d{4}$"): continue

                fecha_w = [w for w in wl if 35 <= w["x0"] < 140]
                conc_w  = [w for w in wl if 140 <= w["x0"] < cargo_x - 4]
                carg_w  = [w for w in wl if cargo_x - 4 <= w["x0"] < abono_x - 4]
                abon_w  = [w for w in wl if abono_x - 4 <= w["x0"] < abono_x + 80]

                fecha = _txt(fecha_w)
                desc  = _txt(conc_w).strip()
                cargo = parse_anglosajón("".join(w["text"] for w in carg_w))
                abono = parse_anglosajón("".join(w["text"] for w in abon_w))

                if cargo > 0:   valor, tipo = -cargo, "DEBITO"
                elif abono > 0: valor, tipo = abono, "CREDITO"
                else: continue
                rows.append({"FECHA": fecha, "DESCRIPCION": desc,
                              "VALOR": valor, "TIPO": tipo, "GMF": is_gmf(desc)})
    return pd.DataFrame(rows) if rows else _empty_df()

# -- COLPATRIA ------------------------------------------------
# Estructura: FECHA | OFICINA | No DOCUM | DESCRIPCION | MONTO | SALDO
# Fecha: d/mm/yyyy   Números: europeo (punto=miles, coma=decimal)
# TX puede ocupar 2 líneas (segunda línea = más descripción sin monto)

def extract_colpatria(pdf_file, pdf_password=""):
    rows = []
    with pdfplumber.open(pdf_file, password=pdf_password or None) as pdf:
        for page in pdf.pages:
            lns = _lines(page)
            monto_x = saldo_x = None
            for y, wl in lns.items():
                t = _txt(wl).upper()
                if "MONTO" in t and "SALDO" in t:
                    for w in wl:
                        if "MONTO" in w["text"].upper(): monto_x = w["x0"]
                        if "SALDO" in w["text"].upper(): saldo_x = w["x0"]
                    break
            if monto_x is None: monto_x, saldo_x = 390, 460

            pending = None
            for y, wl in lns.items():
                if not wl: continue
                first = wl[0]["text"]
                is_tx = bool(re.match(r"^\d{1,2}/\d{2}/\d{4}$", first))

                if is_tx:
                    if pending: rows.append(pending)
                    desc_w  = [w for w in wl if 65 <= w["x0"] < monto_x - 4]
                    monto_w = [w for w in wl if monto_x - 4 <= w["x0"] < saldo_x - 4]
                    desc  = _txt(desc_w).strip()
                    mstr  = "".join(w["text"] for w in monto_w).strip()
                    monto = parse_europeo(mstr)
                    if monto == 0 and not mstr: pending = None; continue
                    tipo = "CREDITO" if monto > 0 else "DEBITO"
                    pending = {"FECHA": first, "DESCRIPCION": desc,
                               "VALOR": monto, "TIPO": tipo, "GMF": is_gmf(desc)}
                elif pending is not None:
                    extra = _txt(wl).strip()
                    if extra and not re.search(r"\d{4,}", extra.replace(".", "").replace(",", "")):
                        pending["DESCRIPCION"] = (pending["DESCRIPCION"] + " " + extra).strip()
                        pending["GMF"] = is_gmf(pending["DESCRIPCION"])
            if pending: rows.append(pending)
    return pd.DataFrame(rows) if rows else _empty_df()

# -- DAVIVIENDA -----------------------------------------------
# Estructura: Día | Mes | Oficina | Descripción | Doc. | Débito | Crédito | Saldo
# TX: empieza con 2 dígitos (día) + 2 dígitos (mes)

def extract_davivienda(pdf_file, pdf_password=""):
    rows = []
    with pdfplumber.open(pdf_file, password=pdf_password or None) as pdf:
        for page in pdf.pages:
            lns = _lines(page)
            deb_x = cred_x = None
            for y, wl in lns.items():
                t = _txt(wl).upper()
                if ("DÉBIT" in t or "DEBIT" in t) and ("CRÉDIT" in t or "CREDIT" in t):
                    for w in wl:
                        if "BIT" in w["text"].upper(): deb_x = w["x0"]
                        if "DIT" in w["text"].upper():  cred_x = w["x0"]
                    break
            if deb_x is None: deb_x, cred_x = 380, 460

            pending = None
            for y, wl in lns.items():
                if len(wl) < 3: continue
                f, s = wl[0]["text"], wl[1]["text"]
                is_tx = bool(re.match(r"^\d{2}$", f) and re.match(r"^\d{2}$", s))
                if is_tx:
                    if pending: rows.append(pending)
                    fecha = f"{f}/{s}"
                    desc_w = [w for w in wl if 55 <= w["x0"] < deb_x - 4]
                    deb_w  = [w for w in wl if deb_x - 4 <= w["x0"] < cred_x - 4]
                    cred_w = [w for w in wl if cred_x - 4 <= w["x0"] < cred_x + 80]
                    desc  = _txt(desc_w).strip()
                    deb   = parse_anglosajón("".join(w["text"] for w in deb_w))
                    cred  = parse_anglosajón("".join(w["text"] for w in cred_w))
                    if deb > 0.01:   valor, tipo = -deb, "DEBITO"
                    elif cred > 0.01: valor, tipo = cred, "CREDITO"
                    else: pending = None; continue
                    pending = {"FECHA": fecha, "DESCRIPCION": desc,
                               "VALOR": valor, "TIPO": tipo, "GMF": is_gmf(desc)}
                elif pending:
                    extra = _txt(wl).strip()
                    if extra and not re.match(r"^\$[\d,\.]+", extra):
                        pending["DESCRIPCION"] = (pending["DESCRIPCION"] + " " + extra).strip()
                        pending["GMF"] = is_gmf(pending["DESCRIPCION"])
            if pending: rows.append(pending)
    return pd.DataFrame(rows) if rows else _empty_df()

# -- OCCIDENTE ------------------------------------------------
# Estructura: DIA | TRANSACCIÓN | IDENT. | DEBITOS | CREDITOS | SALDO
# TX: empieza con día de 2 dígitos

def extract_occidente(pdf_file, pdf_password=""):
    rows = []
    with pdfplumber.open(pdf_file, password=pdf_password or None) as pdf:
        for page in pdf.pages:
            lns = _lines(page)
            deb_x = cred_x = None
            for y, wl in lns.items():
                t = _txt(wl).upper()
                if "DEBITO" in t and "CREDITO" in t:
                    for w in wl:
                        if "DEBIT" in w["text"].upper(): deb_x = w["x0"]
                        if "CREDIT" in w["text"].upper(): cred_x = w["x0"]
                    break
            if deb_x is None: deb_x, cred_x = 340, 430

            for y, wl in lns.items():
                if not wl: continue
                if not _match(wl[0]["text"], r"^\d{2}$"): continue
                fecha = wl[0]["text"]
                desc_w = [w for w in wl if 18 <= w["x0"] < deb_x - 4]
                deb_w  = [w for w in wl if deb_x - 4 <= w["x0"] < cred_x - 4]
                cred_w = [w for w in wl if cred_x - 4 <= w["x0"] < cred_x + 80]
                desc  = _txt(desc_w).strip()
                deb   = parse_anglosajón("".join(w["text"] for w in deb_w))
                cred  = parse_anglosajón("".join(w["text"] for w in cred_w))
                if deb > 0.01:   valor, tipo = -deb, "DEBITO"
                elif cred > 0.01: valor, tipo = cred, "CREDITO"
                else: continue
                rows.append({"FECHA": fecha, "DESCRIPCION": desc,
                              "VALOR": valor, "TIPO": tipo, "GMF": is_gmf(desc)})
    return pd.DataFrame(rows) if rows else _empty_df()

# -- IRIS -----------------------------------------------------
# Estructura: DÍA | REFERENCIA | DESCRIPCIÓN | MOVIMIENTOS | SALDO
# Fecha: dd/mm/yy   MOVIMIENTOS: $ negativo=débito, positivo=crédito

def extract_iris(pdf_file, pdf_password=""):
    rows = []
    with pdfplumber.open(pdf_file, password=pdf_password or None) as pdf:
        for page in pdf.pages:
            lns = _lines(page)
            movi_x = None
            for y, wl in lns.items():
                if "MOVIMIENTO" in _txt(wl).upper():
                    for w in wl:
                        if "MOVI" in w["text"].upper(): movi_x = w["x0"]
                    break
            if movi_x is None: movi_x = 380

            for y, wl in lns.items():
                if not wl: continue
                if not _match(wl[0]["text"], r"^\d{2}/\d{2}/\d{2,4}$"): continue
                fecha = wl[0]["text"]
                desc_w = [w for w in wl if 105 <= w["x0"] < movi_x - 4]
                movi_w = [w for w in wl if movi_x - 4 <= w["x0"] < movi_x + 100]
                desc  = _txt(desc_w).strip()
                mstr  = "".join(w["text"] for w in movi_w).replace("$", "").replace(" ", "")
                valor = parse_anglosajón(mstr)
                if valor == 0: continue
                rows.append({"FECHA": fecha, "DESCRIPCION": desc,
                              "VALOR": valor, "TIPO": "CREDITO" if valor > 0 else "DEBITO",
                              "GMF": is_gmf(desc)})
    return pd.DataFrame(rows) if rows else _empty_df()

def extract_finandina(pdf_file, pdf_password=""):
    st.warning("Extractor Finandina pendiente. Adjunta un extracto para calibrarlo.")
    return _empty_df()

EXTRACTORS = {
    "Bancolombia": extract_bancolombia,
    "BBVA":        extract_bbva,
    "Colpatria":   extract_colpatria,
    "Davivienda":  extract_davivienda,
    "Occidente":   extract_occidente,
    "IRIS":        extract_iris,
    "Finandina":   extract_finandina,
}

# ==============================================================
# EXTRACTOR DMS  (MOVIMIENTOS BANCO MARZO XXXX.xlsx)
# ==============================================================
# Estructura (14 columnas):
#   0=Doc/Pref+Num  1=Descripción(fecha+nombre+notas)  ...
#   5=Débito DMS (dinero entrando al banco = CREDITO en extracto)
#   6=Crédito DMS (dinero saliendo del banco = DEBITO en extracto)
# Fila 0 = encabezado, Fila 1 = resumen de cuenta (se omiten).

_MONTH_MAP = {
    "jan":"01","feb":"02","mar":"03","apr":"04","may":"05","jun":"06",
    "jul":"07","aug":"08","sep":"09","oct":"10","nov":"11","dec":"12",
    "ene":"01","abr":"04","ago":"08",
}

def _parse_dms_fecha(desc_str):
    """Extrae fecha del inicio de la descripción DMS: '02-Mar-2026 Notas Mov:...'"""
    m = re.match(r"^(\d{1,2})-(\w{3})-(\d{4})", str(desc_str).strip())
    if m:
        d, mon, y = m.group(1).zfill(2), m.group(2).lower()[:3], m.group(3)
        return f"{d}/{_MONTH_MAP.get(mon, '??')}"
    return ""

def _parse_dms_desc(desc_str):
    """Extrae nombre y nota de la descripción DMS."""
    s = str(desc_str)
    nombre = nota = ""
    if "Notas Mov:" in s:
        after = s.split("Notas Mov:")[1]
        if "Notas Doc:" in after:
            nombre = after.split("Notas Doc:")[0].strip()
            nota   = after.split("Notas Doc:")[1].strip()
        else:
            nombre = after.strip()
    return (nombre + (" – " + nota if nota and nota not in ("", "None", "Nulo") else "")).strip(" –")

def _doc_valido(x) -> bool:
    """Descarta filas de resumen (doc='nan') y filas vacías del DMS."""
    s = str(x).strip()
    if s in ("nan", "None", "", "0"): return False
    return bool(re.match(r"[A-Za-z0-9]", s))

def extract_dms(excel_file) -> pd.DataFrame:
    """
    Extrae movimientos del Excel DMS (MOVIMIENTOS BANCO MARZO XXXX.xlsx).

    Estructura fija de 14 columnas:
      col[0]  = Documento (ref: '40 59528', 'RCX 14334', '90 66610'…)
      col[1]  = Descripción: 'DD-Mon-YYYY Notas Mov: NOMBRE Notas Doc: NOTA'
      col[5]  = Débito DMS  → dinero ENTRANDO al banco  (CREDITO en extracto)
      col[6]  = Crédito DMS → dinero SALIENDO del banco (DEBITO en extracto)

    Regla 1-a-1:  si col5>0 y col6==0  → CREDITO
                  si col6>0 y col5==0  → DEBITO
                  si ambos o ninguno   → ignorar

    Nota: el archivo tiene una fila de resumen de apertura (fila 1, Cuenta=número)
    y una fila de cierre (doc=NaN) con los totales que se descartan por _doc_valido().
    """
    try:
        df_raw = pd.read_excel(excel_file, header=None, dtype=str, sheet_name=0)
        # Fila 0 = encabezado, Fila 1 = resumen apertura → empezar en fila 2
        rows = []
        for _, row in df_raw.iloc[2:].iterrows():
            doc = str(row.iloc[0]).strip()

            # Descartar fila de cierre (doc='nan') y filas vacías
            if not _doc_valido(doc):
                continue

            desc  = str(row.iloc[1]).strip()
            deb5  = str(row.iloc[5]).strip() if len(row) > 5 else "nan"
            cred6 = str(row.iloc[6]).strip() if len(row) > 6 else "nan"

            fecha       = _parse_dms_fecha(desc)
            descripcion = _parse_dms_desc(desc)
            if not descripcion:
                descripcion = doc

            # col5 = Débito DMS = dinero entrando al banco = CREDITO en extracto
            # col6 = Crédito DMS = dinero saliendo del banco = DEBITO en extracto
            v5 = 0.0 if deb5  in ("nan", "None", "", "0", "0.0")       else parse_anglosajón(deb5)
            v6 = 0.0 if cred6 in ("nan", "None", "", "0", "0.0", "M")  else parse_anglosajón(cred6)

            # Solo usar filas donde EXACTAMENTE UNO de los dos tiene valor
            if v5 > 0 and v6 == 0:
                valor, tipo = v5, "CREDITO"
            elif v6 > 0 and v5 == 0:
                valor, tipo = -v6, "DEBITO"
            else:
                continue  # ambos > 0 o ninguno → ignorar

            if abs(valor) < 1:
                continue

            rows.append({
                "FECHA":       fecha,
                "DESCRIPCION": descripcion,
                "VALOR":       valor,
                "TIPO":        tipo,
                "DOC_DMS":     re.sub(r"\s+", " ", doc).strip(),
                "GMF":         is_gmf(descripcion),
            })
        return pd.DataFrame(rows) if rows else _empty_df()
    except Exception as e:
        st.error(f"Error leyendo DMS: {e}")
        st.exception(e)
        return _empty_df()

# ==============================================================
# CONCILIACIÓN
# ==============================================================

def _find_subset(d_avail: pd.DataFrame, objetivo: int, max_n: int = 15, tol: int = 1):
    """
    Busca subconjunto de filas en d_avail cuya suma de VALOR.abs() == objetivo ± tol.
    Retorna lista de índices del DataFrame original, o None si no encuentra.
    Usado para cruzar N registros DMS contra 1 movimiento bancario.
    """
    from itertools import combinations

    cands = d_avail[d_avail["VALOR"].abs() <= objetivo + tol + 1].copy()
    if len(cands) < 2:
        return None

    vals = cands["VALOR"].abs().round(0).astype(int).tolist()
    idxs = cands.index.tolist()
    pairs = sorted(zip(vals, idxs), reverse=True)[:25]
    vals_s = [v for v, _ in pairs]
    idxs_s = [i for _, i in pairs]

    # Límites de pool por n para evitar explosión combinatoria
    limits = {2: 25, 3: 25, 4: 20, 5: 15, 6: 12, 7: 10, 8: 8, 9: 7, 10: 6}

    for n in range(2, min(max_n + 1, len(pairs) + 1)):
        if sum(vals_s[:n]) < objetivo - tol:
            continue
        pool = min(len(pairs), limits.get(n, 6))
        for combo in combinations(range(pool), n):
            if abs(sum(vals_s[i] for i in combo) - objetivo) <= tol:
                return [idxs_s[i] for i in combo]

    return None


def _resultado_vacio(df_banco):
    """Resultado cuando el DMS está vacío - todo el banco queda sin cruzar."""
    gmf_col = "GMF" in df_banco.columns
    gmf  = df_banco[df_banco["GMF"]].copy()  if gmf_col else _empty_df()
    solo = df_banco[~df_banco["GMF"]].copy() if gmf_col else df_banco.copy()
    return {"mas_df": _empty_df(), "menos_df": _empty_df(),
            "solo_bco": solo, "x_rev_df": _empty_df(), "gmf_df": gmf}

def conciliar(df_banco: pd.DataFrame, df_dms: pd.DataFrame) -> dict:
    """
    Cruce 1-a-1 por valor absoluto (tolerancia ±1 peso).
    Separa CREDITO y DEBITO antes de cruzar.
    GMF excluido del cruce principal.

    Retorna:
      mas_df    → cruzados CREDITOS  (hoja "+")
      menos_df  → cruzados DEBITOS   (hoja "-")
      x_rev_df  → DMS sin match      (hoja "X REVISAR")
      solo_bco  → Banco sin match    (para carátula secciones a/b)
      gmf_df    → GMF banco          (hoja "GMF")
    """
    # Guard: DMS vacío → todo el banco queda sin cruzar
    if df_dms.empty or "TIPO" not in df_dms.columns:
        return _resultado_vacio(df_banco)

    banco_op = df_banco[df_banco["GMF"] == False].copy().reset_index(drop=True)
    dms_op   = df_dms[df_dms["GMF"] == False].copy().reset_index(drop=True)

    mas_rows, menos_rows = [], []
    solo_banco_rows, solo_dms_rows = [], []

    for tipo in ["CREDITO", "DEBITO"]:
        b = banco_op[banco_op["TIPO"] == tipo].copy().reset_index(drop=True)
        d = dms_op[dms_op["TIPO"] == tipo].copy().reset_index(drop=True)

        b["_KEY"] = b["VALOR"].abs().round(0).astype(int)
        d["_KEY"] = d["VALOR"].abs().round(0).astype(int)

        used_d = set()
        solo_b_tipo = []

        # Paso 1: cruce 1-a-1
        for bi, b_row in b.iterrows():
            bk = b_row["_KEY"]
            cands = d[(d["_KEY"].between(bk - 1, bk + 1)) & (~d.index.isin(used_d))]
            if not cands.empty:
                best = cands.iloc[0]
                dif  = abs(b_row["VALOR"]) - abs(best["VALOR"])
                matched = {
                    "FECHA":       b_row["FECHA"],
                    "DESCRIPCION": b_row["DESCRIPCION"],
                    "VALOR":       b_row["VALOR"],
                    "Valor DMS":   best["VALOR"],
                    "Doc Dms":     best.get("DOC_DMS", ""),
                    "# DMS":       1,
                    "Dif":         round(dif, 2),
                }
                if tipo == "CREDITO": mas_rows.append(matched)
                else:                 menos_rows.append(matched)
                used_d.add(cands.index[0])
            else:
                solo_b_tipo.append(b_row)

        # Paso 2: cruce N-DMS-a-1-Banco (un ingreso bancario = varios recibos DMS)
        final_solo_b = []
        for b_row in solo_b_tipo:
            objetivo  = int(round(abs(b_row["VALOR"])))
            d_avail   = d[~d.index.isin(used_d)]
            found     = _find_subset(d_avail, objetivo)
            if found is not None:
                matched_dms = d.loc[found]
                dms_sum = matched_dms["VALOR"].abs().sum()
                docs = [str(r.get("DOC_DMS", "")) for _, r in matched_dms.iterrows()
                        if str(r.get("DOC_DMS", "")).strip() not in ("", "nan", "None")]
                matched = {
                    "FECHA":       b_row["FECHA"],
                    "DESCRIPCION": b_row["DESCRIPCION"],
                    "VALOR":       b_row["VALOR"],
                    "Valor DMS":   dms_sum if tipo == "CREDITO" else -dms_sum,
                    "Doc Dms":     ", ".join(docs),
                    "# DMS":       len(found),
                    "Dif":         round(abs(b_row["VALOR"]) - dms_sum, 2),
                }
                if tipo == "CREDITO": mas_rows.append(matched)
                else:                 menos_rows.append(matched)
                for idx in found:
                    used_d.add(idx)
            else:
                final_solo_b.append(b_row)

        solo_banco_rows.extend(final_solo_b)

        for di, d_row in d.iterrows():
            if di not in used_d:
                solo_dms_rows.append(d_row)

    def to_df(lst, cols=None):
        if not lst: return pd.DataFrame(columns=cols or [])
        return pd.DataFrame(lst).drop(columns=["_KEY"], errors="ignore").reset_index(drop=True)

    match_cols = ["FECHA", "DESCRIPCION", "VALOR", "Valor DMS", "Doc Dms", "# DMS", "Dif"]
    banco_cols  = ["FECHA", "DESCRIPCION", "VALOR", "TIPO", "GMF"]

    return {
        "mas_df":    to_df(mas_rows,         match_cols),
        "menos_df":  to_df(menos_rows,       match_cols),
        "solo_bco":  to_df(solo_banco_rows,  banco_cols),
        "x_rev_df":  to_df(solo_dms_rows),   # DMS sin match
        "gmf_df":    df_banco[df_banco["GMF"]].copy(),
    }

# ==============================================================
# GENERADOR EXCEL  (7 hojas)
# ==============================================================

# - Estilos --------------------------------------------------
def _fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _border_thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

FILL_AZUL   = _fill("1F4E79")
FILL_VERDE  = _fill("1E8449")
FILL_ROJO   = _fill("C0392B")
FILL_NARANJA= _fill("E67E22")
FILL_ALT    = _fill("EBF5FB")
FILL_ALT2   = _fill("FDFEFE")
FILL_AGRUP  = _fill("D6EAF8")   # azul suave – ingreso dividido en varios DMS
FONT_HDR    = _font(bold=True, color="FFFFFF", size=10)
FONT_BOLD   = _font(bold=True, size=10)
FONT_NORM   = _font(size=10)

def _col_letter(n):
    r = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        r = chr(65 + rem) + r
    return r

def _write_df(ws, df, start_row=1, hdr_fill=None, num_cols=None, mark_col=None, mark_fill=None):
    """Escribe DataFrame en hoja con header coloreado y filas alternadas.
    Si mark_col se indica, resalta con mark_fill las filas donde ese campo > 1."""
    if hdr_fill is None: hdr_fill = FILL_AZUL
    if num_cols is None: num_cols = []
    cols = list(df.columns)

    for j, col in enumerate(cols, 1):
        c = ws.cell(start_row, j, col)
        c.fill = hdr_fill; c.font = FONT_HDR
        c.alignment = Alignment(horizontal="center", vertical="center")

    for i, (_, row) in enumerate(df.iterrows(), 1):
        es_agrupado = mark_col and mark_col in df.columns and (row.get(mark_col, 1) or 1) > 1
        fill = (mark_fill or FILL_AGRUP) if es_agrupado else (FILL_ALT if i % 2 == 0 else FILL_ALT2)
        for j, col in enumerate(cols, 1):
            val = row[col]
            if hasattr(val, "item"): val = val.item()
            if isinstance(val, float) and str(val) == "nan": val = ""
            c = ws.cell(start_row + i, j, val)
            c.fill = fill; c.font = FONT_NORM
            if col in num_cols or col in ("VALOR", "Valor DMS", "Dif"):
                c.number_format = "#,##0.00"

    for j, col in enumerate(cols, 1):
        max_len = max(len(str(col)), max(
            (len(str(df.iloc[i][col])) for i in range(min(len(df), 100))), default=0
        ))
        ws.column_dimensions[_col_letter(j)].width = min(max_len + 3, 55)

# - Hoja CARATULA --------------------------------------------

def _sc(ws, cell, val, bold=False, size=10, num_fmt=None, color="000000", align="left"):
    if cell not in (None, ""):
        ws[cell] = val
        ws[cell].font = Font(bold=bold, color=color, size=size, name="Calibri")
        ws[cell].alignment = Alignment(horizontal=align)
        if num_fmt: ws[cell].number_format = num_fmt

def _fill_seccion(ws, start_row, df_subset, max_rows):
    """Llena filas de una sección de la carátula (Fecha, Concepto, Importe, Observaciones, Doc DMS)"""
    for i, (_, row) in enumerate(df_subset.head(max_rows).iterrows()):
        r = start_row + i
        ws.cell(r, 2, str(row.get("FECHA", ""))).font = FONT_NORM
        desc = str(row.get("DESCRIPCION", row.get("DESCRIPCION", "")))
        ws.cell(r, 3, desc).font = FONT_NORM
        val = abs(row.get("VALOR", 0))
        c = ws.cell(r, 4, val)
        c.font = FONT_NORM; c.number_format = "#,##0.00"
        # Doc DMS si existe
        if "DOC_DMS" in row:
            ws.cell(r, 6, str(row.get("DOC_DMS", ""))).font = FONT_NORM

def _write_caratula(ws, df_banco, df_dms, resultado,
                    nombre_banco, cuenta_contable, num_cuenta,
                    saldo_banco, saldo_dms, elaborado_por, fecha_corte):
    ws.sheet_view.showGridLines = False

    # -- Encabezado --
    _sc(ws, "B1", "Conciliación Bancaria", bold=True, size=14, align="center")
    ws.merge_cells("B1:F1")
    ws["B1"].fill = FILL_AZUL; ws["B1"].font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    _sc(ws, "B2", "Empresa:");  _sc(ws, "C2", "SEIKOU SA", bold=True)
    _sc(ws, "E2", "Fecha");     _sc(ws, "F2", str(fecha_corte))
    _sc(ws, "B3", "Banco:");    _sc(ws, "C3", f"BANCO {nombre_banco.upper()}", bold=True)
    _sc(ws, "E3", "Saldo Extracto Banco")
    ws["F3"] = saldo_banco; ws["F3"].number_format = "#,##0.00"; ws["F3"].font = FONT_BOLD
    _sc(ws, "B4", "Cta Cont."); _sc(ws, "C4", cuenta_contable)
    _sc(ws, "E4", "Saldo Cta. Contable")
    ws["F4"] = saldo_dms; ws["F4"].number_format = "#,##0.00"; ws["F4"].font = FONT_BOLD
    _sc(ws, "B5", "# Cta");    _sc(ws, "C5", num_cuenta)

    # -- Sumas por sección --
    sb = resultado["solo_bco"]
    xr = resultado["x_rev_df"]

    a = abs(sb[sb["TIPO"] == "DEBITO"]["VALOR"].sum()) if not sb.empty else 0
    b = abs(sb[sb["TIPO"] == "CREDITO"]["VALOR"].sum()) if not sb.empty else 0
    if not xr.empty and "TIPO" in xr.columns:
        c = abs(xr[xr["TIPO"] == "DEBITO"]["VALOR"].sum())
        d = abs(xr[xr["TIPO"] == "CREDITO"]["VALOR"].sum())
    else:
        c = d = 0
    formula = a - b - c + d
    diferencia = round(saldo_banco - saldo_dms, 2)

    _sc(ws, "B7", "a) Pagos Banco no Contabilidad");    ws["D7"] = a; ws["D7"].number_format = "#,##0.00"
    _sc(ws, "E7", "a-b-c+d");                           ws["F7"] = formula; ws["F7"].number_format = "#,##0.00"
    _sc(ws, "B8", "b) Cobros Banco no Contabilidad");   ws["D8"] = b; ws["D8"].number_format = "#,##0.00"
    _sc(ws, "B9", "c) Pagos  Contabilidad no Banco");   ws["D9"] = c; ws["D9"].number_format = "#,##0.00"
    _sc(ws, "E9", "Diferencia");                         ws["F9"] = diferencia; ws["F9"].number_format = "#,##0.00"
    _sc(ws, "B10", "d) Cobros  Contabilidad no Banco"); ws["D10"] = d; ws["D10"].number_format = "#,##0.00"

    # color diferencia
    dif_color = "00AA00" if abs(diferencia) < 1 else "FF0000"
    ws["F9"].font = Font(bold=True, color=dif_color, size=10, name="Calibri")

    # -- Sección a) Pagos Banco no Contabilidad --
    _sc(ws, "B12", "a) Pagos Banco no Contabilidad  REVISAR TESORERIA", bold=True, size=10)
    ws["B12"].fill = _fill("D6E4F7")
    _sc(ws, "B14", "Fecha"); _sc(ws, "C14", "Concepto")
    _sc(ws, "D14", "Importe"); _sc(ws, "E14", "Observaciones"); _sc(ws, "F14", "Clave Conciliación")
    for col in ["B14","C14","D14","E14","F14"]:
        ws[col].font = FONT_HDR; ws[col].fill = FILL_AZUL
    sb_deb = sb[sb["TIPO"] == "DEBITO"] if not sb.empty else pd.DataFrame()
    _fill_seccion(ws, 15, sb_deb, 5)
    ws["C20"] = "Total"; ws["C20"].font = FONT_BOLD
    ws["D20"] = a; ws["D20"].number_format = "#,##0.00"; ws["D20"].font = FONT_BOLD

    # -- Sección b) Cobros Banco no Contabilidad --
    _sc(ws, "B22", "b) Consignaciones  Banco no Contabilidad", bold=True)
    ws["B22"].fill = _fill("D6E4F7")
    for col in ["B24","C24","D24","E24","F24"]:
        ws[col].font = FONT_HDR; ws[col].fill = FILL_AZUL
    _sc(ws, "B24", "Fecha"); _sc(ws, "C24", "Concepto")
    _sc(ws, "D24", "Importe"); _sc(ws, "E24", "Observaciones"); _sc(ws, "F24", "Clave Conciliación")
    sb_cred = sb[sb["TIPO"] == "CREDITO"] if not sb.empty else pd.DataFrame()
    _fill_seccion(ws, 25, sb_cred, 6)
    ws["C31"] = "Total"; ws["C31"].font = FONT_BOLD
    ws["D31"] = b; ws["D31"].number_format = "#,##0.00"; ws["D31"].font = FONT_BOLD

    # -- Sección c) Pagos Contabilidad no Banco --
    _sc(ws, "B33", "c) Pagos - notas Contabilidad no Banco  REVISAR TESORERIA", bold=True)
    ws["B33"].fill = _fill("D6E4F7")
    for col in ["B35","C35","D35","E35","F35"]:
        ws[col].font = FONT_HDR; ws[col].fill = FILL_AZUL
    _sc(ws, "B35", "Fecha"); _sc(ws, "C35", "Concepto")
    _sc(ws, "D35", "Importe"); _sc(ws, "E35", "Observaciones"); _sc(ws, "F35", "Clave Conciliación")
    xr_deb = xr[xr["TIPO"] == "DEBITO"] if not xr.empty and "TIPO" in xr.columns else pd.DataFrame()
    _fill_seccion(ws, 36, xr_deb, 10)
    ws["C46"] = "Total"; ws["C46"].font = FONT_BOLD
    ws["D46"] = c; ws["D46"].number_format = "#,##0.00"; ws["D46"].font = FONT_BOLD

    # -- Sección d) Cobros Contabilidad no Banco --
    _sc(ws, "B48", "d) Consignaciones Contabilidad no Banco  TRANSICION", bold=True)
    ws["B48"].fill = _fill("D6E4F7")
    for col in ["B50","C50","D50","E50","F50"]:
        ws[col].font = FONT_HDR; ws[col].fill = FILL_AZUL
    _sc(ws, "B50", "Fecha"); _sc(ws, "C50", "Concepto")
    _sc(ws, "D50", "Importe"); _sc(ws, "E50", "Observaciones"); _sc(ws, "F50", "Clave Conciliación")
    xr_cred = xr[xr["TIPO"] == "CREDITO"] if not xr.empty and "TIPO" in xr.columns else pd.DataFrame()
    _fill_seccion(ws, 51, xr_cred, 8)
    ws["C59"] = "Total"; ws["C59"].font = FONT_BOLD
    ws["D59"] = d; ws["D59"].number_format = "#,##0.00"; ws["D59"].font = FONT_BOLD

    # -- Firma --
    ws["B63"] = "ELABORO"; ws["B63"].font = FONT_BOLD
    ws["B64"] = elaborado_por; ws["B64"].font = FONT_NORM

    # Anchos
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 20


def _write_leyenda_agrup(ws, row):
    """Escribe una leyenda de color azul debajo de los datos de cruce."""
    c_cuad = ws.cell(row, 1, "")
    c_cuad.fill = FILL_AGRUP
    c_txt = ws.cell(row, 2, "Fondo azul = ingreso bancario agrupado en varios recibos DMS (ver columna # DMS)")
    c_txt.font = Font(italic=True, color="1A5276", size=9, name="Calibri")


def generar_excel(df_banco, df_dms, resultado,
                  nombre_banco, cuenta_contable, num_cuenta,
                  saldo_banco, saldo_dms, elaborado_por, fecha_corte,
                  plantilla_file=None):

    if plantilla_file:
        wb = openpyxl.load_workbook(plantilla_file)
    else:
        wb = openpyxl.Workbook()
        for name in ["CARATULA", "BANCO", "DMS", "+", "-", "X REVISAR", "GMF"]:
            wb.create_sheet(name)
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    def _get(name):
        if name not in wb.sheetnames: wb.create_sheet(name)
        ws = wb[name]
        if ws.max_row > 0: ws.delete_rows(1, ws.max_row + 1)
        return ws

    # -- CARATULA --
    ws_car = _get("CARATULA")
    _write_caratula(ws_car, df_banco, df_dms, resultado,
                    nombre_banco, cuenta_contable, num_cuenta,
                    saldo_banco, saldo_dms, elaborado_por, fecha_corte)

    # -- BANCO (FECHA | DESCRIPCIÓN | VALOR) --
    ws_bco = _get("BANCO")
    df_bco_out = df_banco[["FECHA", "DESCRIPCION", "VALOR"]].copy()
    df_bco_out.columns = ["FECHA", "DESCRIPCIÓN", "VALOR"]
    _write_df(ws_bco, df_bco_out, num_cols=["VALOR"])

    # -- DMS (FECHA | DESCRIPCION | VALOR | TIPO | DOC_DMS) --
    ws_dms = _get("DMS")
    dms_cols = ["FECHA", "DESCRIPCION", "VALOR", "TIPO"]
    if "DOC_DMS" in df_dms.columns: dms_cols.append("DOC_DMS")
    _write_df(ws_dms, df_dms[dms_cols], num_cols=["VALOR"])

    # -- + (cruzados CREDITO) --
    ws_mas = _get("+")
    mas = resultado["mas_df"]
    if not mas.empty:
        _write_df(ws_mas, mas, hdr_fill=FILL_VERDE, num_cols=["VALOR","Valor DMS","Dif"],
                  mark_col="# DMS", mark_fill=FILL_AGRUP)
        _write_leyenda_agrup(ws_mas, len(mas) + 3)

    # -- - (cruzados DEBITO) --
    ws_men = _get("-")
    menos = resultado["menos_df"]
    if not menos.empty:
        _write_df(ws_men, menos, hdr_fill=FILL_ROJO, num_cols=["VALOR","Valor DMS","Dif"],
                  mark_col="# DMS", mark_fill=FILL_AGRUP)
        _write_leyenda_agrup(ws_men, len(menos) + 3)

    # -- X REVISAR (DMS sin match) --
    ws_xr = _get("X REVISAR")
    xr = resultado["x_rev_df"]
    if not xr.empty:
        out_cols = ["FECHA", "DESCRIPCION", "TIPO", "VALOR"]
        if "DOC_DMS" in xr.columns:
            out_cols.append("DOC_DMS")
        xr_out = xr[out_cols].copy()
        xr_out["VALOR"] = xr_out["VALOR"].abs()
        _write_df(ws_xr, xr_out, hdr_fill=FILL_NARANJA, num_cols=["VALOR"])

    # -- GMF --
    ws_gmf = _get("GMF")
    gmf = resultado["gmf_df"]
    if not gmf.empty:
        gmf_out = gmf[["FECHA","DESCRIPCION","VALOR"]].copy()
        _write_df(ws_gmf, gmf_out, hdr_fill=_fill("7F8C8D"), num_cols=["VALOR"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ==============================================================
# UI STREAMLIT
# ==============================================================

def _fmt(v): return f"$ {v:>18,.2f}"

def _mostrar(resultado, df_banco, df_dms, banco, saldo_banco, saldo_dms):
    st.divider()
    st.subheader(" Resultados de la conciliación")

    total_b  = len(df_banco)
    total_d  = len(df_dms)
    n_mas    = len(resultado["mas_df"])
    n_men    = len(resultado["menos_df"])
    n_sb     = len(resultado["solo_bco"])
    n_xr     = len(resultado["x_rev_df"])
    n_gmf    = len(resultado["gmf_df"])
    concil   = n_mas + n_men
    pct      = round(concil / max(total_b, 1) * 100, 1)
    dif      = round(saldo_banco - saldo_dms, 2)

    c1,c2,c3,c4,c5,c6 = st.columns(6)
    c1.metric("Mov. Banco",  total_b)
    c2.metric("Mov. DMS",    total_d)
    c3.metric("Conciliados", concil, f"{pct}%")
    c4.metric("Solo Banco",  n_sb,   delta=f"-{n_sb}" if n_sb else None, delta_color="inverse")
    c5.metric("X Revisar",   n_xr,   delta=f"-{n_xr}" if n_xr else None, delta_color="inverse")
    c6.metric("GMF",         n_gmf)

    c7,c8,c9 = st.columns(3)
    c7.metric("Saldo Banco",  _fmt(saldo_banco))
    c8.metric("Saldo DMS",    _fmt(saldo_dms))
    c9.metric("Diferencia",   _fmt(dif),
              delta=" Cuadrado" if abs(dif) < 1 else " Diferencia")

    tabs = st.tabs([
        f" Banco ({total_b})",
        f" DMS ({total_d})",
        f" + Créditos ({n_mas})",
        f" − Débitos ({n_men})",
        f" Solo Banco ({n_sb})",
        f" X Revisar ({n_xr})",
        f" GMF ({n_gmf})",
    ])
    with tabs[0]:
        st.dataframe(df_banco[["FECHA","DESCRIPCION","VALOR","TIPO","GMF"]], use_container_width=True, height=380)
    with tabs[1]:
        cols_dms = ["FECHA","DESCRIPCION","VALOR","TIPO"]
        if "DOC_DMS" in df_dms.columns: cols_dms.append("DOC_DMS")
        st.dataframe(df_dms[cols_dms], use_container_width=True, height=380)
    with tabs[2]:
        if not resultado["mas_df"].empty:
            st.dataframe(resultado["mas_df"], use_container_width=True, height=380)
        else:
            st.info("Sin créditos cruzados")
    with tabs[3]:
        if not resultado["menos_df"].empty:
            st.dataframe(resultado["menos_df"], use_container_width=True, height=380)
        else:
            st.info("Sin débitos cruzados")
    with tabs[4]:
        sb = resultado["solo_bco"]
        if not sb.empty:
            st.dataframe(sb[["FECHA","DESCRIPCION","VALOR","TIPO"]], use_container_width=True, height=380)
        else:
            st.success("¡Todo el banco está en el DMS!")
    with tabs[5]:
        xr = resultado["x_rev_df"]
        if not xr.empty:
            out_cols = ["FECHA", "DESCRIPCION", "TIPO", "VALOR"]
            if "DOC_DMS" in xr.columns:
                out_cols.append("DOC_DMS")
            out = xr[out_cols].copy()
            out["VALOR"] = out["VALOR"].abs()
            st.dataframe(out, use_container_width=True, height=380)
        else:
            st.success("¡Todo el DMS está en el banco!")
    with tabs[6]:
        gmf = resultado["gmf_df"]
        if not gmf.empty:
            st.dataframe(gmf[["FECHA","DESCRIPCION","VALOR"]], use_container_width=True, height=380)
        else:
            st.info("Sin GMF identificados")


def main():
    # -- Sidebar ----------------------------------------------
    with st.sidebar:
        st.title(" Configuración")
        banco = st.selectbox("Banco", BANCOS)

        info = INFO_BANCO.get(banco, {})
        cuenta_contable = st.text_input("Cuenta Contable DMS", info.get("cuenta", ""))
        num_cuenta      = st.text_input("N° Cuenta Bancaria",   info.get("num_cta", ""))
        fecha_corte     = st.date_input("Fecha de Corte",        date.today())
        elaborado_por   = st.text_input("Elaborado por",         "Contabilidad Seikou")

        st.divider()
        st.subheader(" Saldos")
        saldo_banco = st.number_input("Saldo Extracto Banco ($)", value=0.0, format="%.2f", step=1000.0)
        saldo_dms   = st.number_input("Saldo Contable DMS ($)",   value=0.0, format="%.2f", step=1000.0)

        st.divider()
        plantilla_file = st.file_uploader(" Plantilla Excel (opcional)", type=["xlsx"])

        pdf_password = st.text_input("Contrasena PDF (si aplica)", value="", type="password",
                                     help="Algunos bancos protegen el extracto con el NIT de la empresa")
        st.divider()
        debug_mode = st.checkbox(" Debug: ver posiciones PDF")

    # -- Área principal ----------------------------------------
    st.title(" Conciliación Bancaria - Seikou S.A.")

    col1, col2 = st.columns(2)
    with col1:
        st.subheader(f" Extracto bancario – {banco}")
        pdf_file = st.file_uploader(f"Subir PDF {banco}", type=["pdf"], key="pdf_up")
    with col2:
        st.subheader(" Movimientos DMS")
        excel_file = st.file_uploader(
            f"Subir Excel DMS  (MOVIMIENTOS {banco.upper()} MARZO…xlsx)",
            type=["xlsx", "xls"], key="dms_up"
        )

    # Debug: muestra palabras del PDF con posición x
    if debug_mode and pdf_file:
        with st.expander(" Primeras 80 palabras del PDF (para calibrar x-posiciones)", expanded=False):
            with pdfplumber.open(pdf_file) as pdf:
                words = pdf.pages[0].extract_words(x_tolerance=3, y_tolerance=5)[:80]
            st.dataframe(pd.DataFrame(words)[["text","x0","top","x1","bottom"]], use_container_width=True)
        pdf_file.seek(0)

    if pdf_file and excel_file:
        if st.button(" Procesar Conciliación", type="primary", use_container_width=True):
            # -- Extraer banco --
            with st.spinner(f"Extrayendo PDF {banco}…"):
                try:
                    import tempfile, os
                    pdf_file.seek(0)
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                        tmp.write(pdf_file.read())
                        tmp_path = tmp.name
                    try:
                        df_banco = EXTRACTORS[banco](tmp_path, pdf_password or "")
                    finally:
                        os.unlink(tmp_path)
                    if df_banco.empty:
                        st.warning(" No se extrajeron movimientos del PDF. Activa modo debug para revisar.")
                    else:
                        st.success(f" {banco}: {len(df_banco)} movimientos  |  "
                                   f"Créditos: {(df_banco.TIPO=='CREDITO').sum()}  "
                                   f"Débitos: {(df_banco.TIPO=='DEBITO').sum()}  "
                                   f"GMF: {df_banco.GMF.sum()}")
                except Exception as e:
                    st.error(f" Error PDF: {e}"); st.exception(e); return

            # -- Extraer DMS --
            with st.spinner("Extrayendo DMS…"):
                try:
                    excel_file.seek(0)
                    df_dms = extract_dms(io.BytesIO(excel_file.read()))
                    if df_dms.empty:
                        # Diagnóstico: mostrar primeras filas del archivo subido
                        excel_file.seek(0)
                        df_diag = pd.read_excel(io.BytesIO(excel_file.read()), header=None, dtype=str, nrows=5)
                        st.warning(" No se extrajeron movimientos del DMS.")
                        st.info(" **Primeras 5 filas del archivo DMS subido** (para diagnóstico):")
                        st.dataframe(df_diag, use_container_width=True)
                        st.info("El archivo DMS debe ser el Excel de **MOVIMIENTOS** "
                                "(ej: *MOVIMIENTOS BANCOLOMBIA MARZO 2026.xlsx*), "
                                "no el extracto bancario ni la plantilla.")
                        return
                    else:
                        st.success(f" DMS: {len(df_dms)} movimientos  |  "
                                   f"Créditos: {(df_dms.TIPO=='CREDITO').sum()}  "
                                   f"Débitos: {(df_dms.TIPO=='DEBITO').sum()}")
                except Exception as e:
                    st.error(f" Error DMS: {e}"); st.exception(e); return

            # -- Conciliar --
            with st.spinner("Conciliando…"):
                resultado = conciliar(df_banco, df_dms)

            # Guardar en session_state
            st.session_state.update({
                "df_banco":  df_banco, "df_dms": df_dms,
                "resultado": resultado, "banco": banco,
                "cfg": dict(banco=banco, cuenta_contable=cuenta_contable,
                            num_cuenta=num_cuenta, saldo_banco=saldo_banco,
                            saldo_dms=saldo_dms, elaborado_por=elaborado_por,
                            fecha_corte=str(fecha_corte)),
            })

            _mostrar(resultado, df_banco, df_dms, banco, saldo_banco, saldo_dms)

            # -- Generar Excel --
            with st.spinner("Generando Excel…"):
                excel_buf = generar_excel(
                    df_banco, df_dms, resultado,
                    banco, cuenta_contable, num_cuenta,
                    saldo_banco, saldo_dms, elaborado_por, str(fecha_corte),
                    plantilla_file,
                )
                st.session_state["excel_buf"] = excel_buf
                st.success(" Excel listo para descargar")

    elif "resultado" in st.session_state:
        cfg = st.session_state["cfg"]
        _mostrar(st.session_state["resultado"],
                 st.session_state["df_banco"],
                 st.session_state["df_dms"],
                 st.session_state.get("banco",""),
                 cfg["saldo_banco"], cfg["saldo_dms"])

    # -- Botón descarga --
    if "excel_buf" in st.session_state:
        st.divider()
        cfg = st.session_state.get("cfg", {})
        fn  = f"conciliacion_{cfg.get('banco','')}_{cfg.get('fecha_corte','')}.xlsx"
        st.download_button(
            " Descargar Excel Conciliación",
            data=st.session_state["excel_buf"],
            file_name=fn,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )

if __name__ == "__main__":
    main()
